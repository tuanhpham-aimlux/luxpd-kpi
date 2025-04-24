import streamlit as st
import shutil
import pandas as pd
import os
import warnings
import tempfile
from datetime import datetime, timedelta
import requests
import base64
import yfinance as yf
from io import BytesIO
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import zipfile

# Suppress openpyxl data validation warning
warnings.filterwarnings("ignore", category=UserWarning, 
                       message="Data Validation extension is not supported and will be removed")

def safe_concat(dataframes):
    """
    Safely concatenate dataframes avoiding the FutureWarning about empty/NA entries
    """
    if not dataframes:
        return pd.DataFrame()
    
    # If only one dataframe, return it directly
    if len(dataframes) == 1:
        return dataframes[0].copy()
    
    # Filter out any completely empty dataframes
    non_empty_dfs = [df for df in dataframes if not df.empty]
    
    if not non_empty_dfs:
        return pd.DataFrame()
    
    # Get all columns that appear in any of the dataframes
    all_columns = set()
    for df in non_empty_dfs:
        all_columns.update(df.columns)
      # Ensure all dataframes have the same columns to avoid warnings
    processed_dfs = []
    for df in non_empty_dfs:
        # Create a copy with all required columns
        new_df = pd.DataFrame()
        for col in all_columns:
            if col in df.columns:
                new_df[col] = df[col]
            else:
                new_df[col] = pd.Series(dtype=object)
        processed_dfs.append(new_df)
    
    # Now concatenate with axis=0 explicitly specified
    if processed_dfs:
        # Filter out any empty or all-NA columns before concatenation to avoid the FutureWarning
        for i, df in enumerate(processed_dfs):
            for col in df.columns:
                if df[col].isna().all():
                    # Replace with a column that has at least one non-NA value
                    df[col] = df[col].copy()
                    if len(df[col]) > 0:
                        df.loc[df.index[0], col] = None  # This ensures column type is preserved but not all-NA
            processed_dfs[i] = df
        
        return pd.concat(processed_dfs, axis=0, ignore_index=True)
    else:
        return pd.DataFrame(columns=list(all_columns))

def get_exchange_rate(date_obj, method="fixed"):
    """
    Get USD/EUR exchange rate for a given date using various methods.
    Compatible with Python 3.12 (no pandas_datareader dependency)
    """
    # Format date as string
    date_str = date_obj.strftime("%Y-%m-%d")
    
    # Fixed rate fallback
    FIXED_RATE = 0.92
    
    try:
        if method == "ecb":
            # European Central Bank historical data
            test_date = date_obj
            max_attempts = 5  # Try up to 5 days back
            attempts = 0
            
            while attempts < max_attempts:
                try:
                    url = f"https://www.ecb.europa.eu/stats/eurofxref/eurofxref-daily.xml"
                    response = requests.get(url)
                    
                    if response.status_code == 200:
                        # Parse XML response
                        root = ET.fromstring(response.content)
                        
                        # Extract USD rate (ECB provides EUR to USD, need to invert)
                        namespaces = {'ns': 'http://www.ecb.int/vocabulary/2002-08-01/eurofxref'}
                        for cube in root.findall('.//ns:Cube[@currency="USD"]', namespaces):
                            usd_eur_rate = 1 / float(cube.attrib['rate'])
                            st.write(f"Retrieved ECB exchange rate for {test_date.strftime('%Y-%m-%d')}: {usd_eur_rate}")
                            return usd_eur_rate
                    
                    # Try previous day
                    test_date -= timedelta(days=1)
                    attempts += 1
                except Exception as e:
                    st.warning(f"Error with ECB API: {e}")
                    test_date -= timedelta(days=1)
                    attempts += 1
            
            st.warning("Could not retrieve ECB data, using fixed rate")
            return FIXED_RATE
            
        elif method == "yahoo":
            # Use Yahoo Finance data directly without pandas_datareader
            import yfinance as yf
            
            # Get data for the previous 5 days in case the exact date isn't available
            end_date = date_obj + timedelta(days=1)
            start_date = date_obj - timedelta(days=5)
            
            try:
                # Get exchange rate data
                ticker_data = yf.download("EURUSD=X", start=start_date, end=end_date, progress=False)
                
                if not ticker_data.empty:
                    # Find the closest date on or before the requested date
                    ticker_data = ticker_data.sort_index()
                    available_dates = ticker_data.index
                    valid_dates = [d for d in available_dates if d <= pd.Timestamp(date_obj)]
                    
                    if valid_dates:
                        closest_date = max(valid_dates)
                        rate = 1 / ticker_data.loc[closest_date, 'Close']
                        st.write(f"Retrieved Yahoo Finance exchange rate for {closest_date.strftime('%Y-%m-%d')}: {rate}")
                        return rate
            except Exception as e:
                st.warning(f"Error with Yahoo Finance: {e}")
            
            st.warning("Could not retrieve Yahoo Finance data, using fixed rate")
            return FIXED_RATE
            
        else:
            # Default to fixed rate
            st.write(f"Using fixed exchange rate: {FIXED_RATE}")
            return FIXED_RATE
            
    except Exception as e:
        st.error(f"Error retrieving exchange rate: {e}")
        st.warning(f"Using fixed exchange rate: {FIXED_RATE}")
        return FIXED_RATE

def read_delta_master_us(file_bytes, asset_manager, year, reporting_date, quarter, usd_eur_fx, date_obj):
    """
    Read and process the US Delta Master file with additional calculated columns.
    """
    try:
        # Read the first sheet with row 1 as headers
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=0, header=0)
        
        # List of columns to extract
        us_columns = [
            'Asset Manager',
            'LOAN_NAME',
            'OFV',
            'CURR_PRIN_BAL',
            'CONSTRUCTION',
            'PROPERTY_TYPE',
            'INTEREST_TYPE',
            'LETTER_RATING',
            'CURR_INT_RATE',
            'CURR_INT_LTV',
            'NRM_DSCR',
            'NRM_NOI',
            'Watchlist',
            'MAT_DATE'
        ]
        
        # Check which required columns exist in the dataframe
        available_columns = [col for col in us_columns if col in df.columns]
        missing_columns = [col for col in us_columns if col not in df.columns]
        
        if missing_columns:
            st.warning(f"The following columns were not found in the US file: {missing_columns}")
            
        if not available_columns:
            st.error("None of the required columns were found in the US file.")
            return pd.DataFrame()
            
        # Filter rows where Asset Manager is in ['ARE', 'JPM']
        df = df[df['Asset Manager'].isin(['ARE', 'JPM'])].copy()
        
        if df.empty:
            st.error("No rows with Asset Manager in ['ARE', 'JPM'] found in the file.")
            return pd.DataFrame()
            
        st.info(f"Filtered to {len(df)} rows with Asset Manager in ['ARE', 'JPM']")
        
        # Clean and transform the data
        
        # 1. Replace 'ARE' with 'PPRE US'
        df['Asset Manager'] = df['Asset Manager'].replace('ARE', 'PPRE US')
        
        # 2. Clean Watchlist column
        df['Watchlist'] = df['Watchlist'].apply(
            lambda x: 'Yes' if pd.notna(x) and str(x).strip() in ['Yellow', 'Red'] else 'No'
        )
        
        df['CURR_INT_LTV'] = df['CURR_INT_LTV'].apply(
            lambda x: x if pd.isna(x) or x == '' else x/100 
        )
        
        # 3. Correct PROPERTY_TYPE values
        property_type_mapping = {
            'Mixed': 'Other',
            'MIXED': 'Other',
            'Various': 'Other',
            'MEDICAL': 'Other',
            'MULTIFAM': 'Residential',
            'WAREHOUSE': 'Logistics',
            'RETAIL': 'Retail',
            'OFFICE': 'Office',
            'Storage': 'Other',
            'STORAGE': 'Other',
            'APARTMENT': 'Residential',
            'Logistic': 'Logistics'
        }
        
        df['PROPERTY_TYPE'] = df['PROPERTY_TYPE'].apply(
            lambda x: property_type_mapping.get(x, x) if pd.notna(x) else x
        )
        
        # 4. Rename columns
        rename_columns = {
            'Asset Manager': 'Asset manager',
            'LOAN_NAME': 'Investment name',
            'OFV': 'Loan commitment (loan CCY)',
            'CURR_PRIN_BAL': 'Loan amount drawn (loan CCY)',
            'CONSTRUCTION': 'CONSTRUCTION',
            'PROPERTY_TYPE': 'Sector',
            'INTEREST_TYPE': 'Fixed vs. Floating',
            'LETTER_RATING': 'Credit Rating',
            'CURR_INT_RATE': 'Current coupon (p.a.)',
            'CURR_INT_LTV': 'Latest LTV',
            'NRM_DSCR': 'DSCR (current) T12 capped',
            'NRM_NOI': 'NRM_NOI',
            'Watchlist': 'Watchlist',
            'MAT_DATE': 'MAT_DATE'
        }
        
        df = df.rename(columns=rename_columns)
        
        # Convert numeric columns
        numeric_columns = [
            'Loan commitment (loan CCY)', 
            'Loan amount drawn (loan CCY)', 
            'Current coupon (p.a.)', 
            'Latest LTV', 
            'DSCR (current) T12 capped',
            'NRM_NOI'
        ]
        
        for col in numeric_columns:
            if col in df.columns:
                # Handle percentage columns
                if col in ['Current coupon (p.a.)', 'Latest LTV']:
                    df[col] = pd.to_numeric(df[col].astype(str).str.replace('%', '').str.replace(',', '.'), errors='coerce') / 100
                else:
                    df[col] = pd.to_numeric(df[col].astype(str).str.replace(',', '.'), errors='coerce')
        
        df['Latest LTV'] = df['Latest LTV'] * 100  # Convert to percentage
        df['Current coupon (p.a.)'] = df['Current coupon (p.a.)'] * 100  # Convert to percentage
        # Add calculated columns
        
        # Basic info columns
        df['Year'] = year
        df['Reporting Date'] = reporting_date
        df['Quarter'] = quarter
        
        # Strategy based on Asset manager
        df['Strategy'] = df['Asset manager'].apply(lambda x: 'PPRE US' if x == 'PPRE US' else 'JPM US')
        
        # Set Loan CCY to USD
        df['Loan CCY'] = 'USD'
        
        # Calculate undrawn amount
        df['Loan amount undrawn/ repaid (loan CCY)'] = df['Loan commitment (loan CCY)'] - df['Loan amount drawn (loan CCY)']
        
        # Convert to EUR using exchange rate
        df['Loan commitment EUR'] = df['Loan commitment (loan CCY)'] * usd_eur_fx
        df['Loan amount drawn EUR'] = df['Loan amount drawn (loan CCY)'] * usd_eur_fx
        df['Loan amount undrawn/ repaid EUR'] = df['Loan amount undrawn/ repaid (loan CCY)'] * usd_eur_fx
        
        # Set Country and Seniority
        df['Country'] = 'US'
        df['Seniority'] = 'Senior'
        
        # Calculate remaining loan terms based on MAT_DATE
        try:
            df['MAT_DATE'] = pd.to_datetime(df['MAT_DATE'])
            df['Remaining loan term w/o extension'] = df['MAT_DATE'].apply(
                lambda x: (x - date_obj).days / 365 if pd.notna(x) else None
            )
            df['Remaining loan term if fully extended'] = df['Remaining loan term w/o extension']
        except Exception as e:
            st.warning(f"Error calculating loan terms: {e}")
            df['Remaining loan term w/o extension'] = None
            df['Remaining loan term if fully extended'] = None
        
        # Empty loan valuation
        df['Loan valuation'] = None        # Weighted calculations
        df['Current coupon (p.a.) weighted'] = df['Current coupon (p.a.)'] * df['Loan amount drawn EUR']
        df['Latest LTV Weighted'] = df['Loan amount drawn EUR'] * df['Latest LTV']
          # LTV range columns
        df['LTV < 55%'] = df['Latest LTV'].apply(lambda x: x if pd.notna(x) and x < 0.55 else None)
        df['55% ≤ LTV < 65%'] = df['Latest LTV'].apply(lambda x: x if pd.notna(x) and 0.55 <= x < 0.65 else None)
        df['65% ≤ LTV < 75%'] = df['Latest LTV'].apply(lambda x: x if pd.notna(x) and 0.65 <= x < 0.75 else None)
        df['75% < LTV'] = df['Latest LTV'].apply(lambda x: x if pd.notna(x) and x >= 0.75 else None)
        
        # DSCR weighted
        df['DSCR (current) Weighted'] = df['Loan amount drawn EUR'] * df['DSCR (current) T12 capped']
        
        # ICR calculation
        df['ICR (current) T12 capped'] = df.apply(
            lambda row: row['NRM_NOI'] / (row['Current coupon (p.a.)'] * row['Loan amount drawn (loan CCY)']) 
            if pd.notna(row['NRM_NOI']) and pd.notna(row['Current coupon (p.a.)']) and 
            pd.notna(row['Loan amount drawn (loan CCY)']) and row['Loan amount drawn (loan CCY)'] > 0 
            and row['Current coupon (p.a.)'] > 0 else None, 
            axis=1
        )
        
        # ICR weighted
        df['ICR (current) Weighted'] = df['Loan commitment EUR'] * df['ICR (current) T12 capped']
        
        # Debt yield calculation
        df['Debt yield (current)'] = df.apply(
            lambda row: row['NRM_NOI'] / row['Loan amount drawn (loan CCY)'] 
            if pd.notna(row['NRM_NOI']) and pd.notna(row['Loan amount drawn (loan CCY)']) 
            and row['Loan amount drawn (loan CCY)'] > 0 else None, 
            axis=1
        )
          # Debt yield weighted
        df['Debt yield (current)'] = df['Debt yield (current)']*100
        df['Debt yield (current) Weighted'] = df['Loan amount drawn EUR'] * df['Debt yield (current)'] 
        
        # Empty impairment
        df['Impairment'] = None
        
        # Development based on CONSTRUCTION
        df['Development'] = df['CONSTRUCTION'].apply(
            lambda x: 'Development' if pd.notna(x) and str(x).strip().upper() == 'Y' else 'Stabilized'
        )
        
        # Identifier BNY Report
        df["Identifier BNY Report 'Issuer Name'"] = None
        
        # Drop temporary columns
        temp_columns = ['CONSTRUCTION', 'NRM_NOI', 'MAT_DATE']
        df = df.drop(columns=[col for col in temp_columns if col in df.columns])
        
        st.success(f"Successfully processed US data with {len(df)} rows")
        return df
        
    except Exception as e:
        st.error(f"Error processing the US file: {e}")
        return pd.DataFrame()

def integrate_watchlist(main_df, watchlist_bytes, sheet_name):
    """
    Process watchlist file and join with main dataframe.
    """
    try:
        # Read the watchlist file, specifying the header row is 7 (zero-based index is 6)
        watchlist_df = pd.read_excel(BytesIO(watchlist_bytes), sheet_name=sheet_name, header=6)
        
        # Select only the first two columns
        if len(watchlist_df.columns) >= 2:
            watchlist_df = watchlist_df.iloc[:, :2].copy()
            
            # Clean up header names (replace \n with spaces)
            watchlist_df.columns = [col.replace('\n', ' ') if isinstance(col, str) else col for col in watchlist_df.columns]
            
            st.write(f"Watchlist columns: {watchlist_df.columns.tolist()}")
            
            # Rename columns to ensure they match expected names
            # Assuming first column is "Deal" and second is "Traffic light"
            if len(watchlist_df.columns) == 2:
                watchlist_df.columns = ["Deal", "Traffic light"]
                
                # Create "Watchlist" column based on "Traffic light" values
                watchlist_df['Watchlist'] = watchlist_df['Traffic light'].apply(
                    lambda x: "No" if x == "Green" else "Yes" if x in ["Red", "Yellow"] else None
                )
                
                # Drop rows with NaN in Deal column
                watchlist_df = watchlist_df.dropna(subset=['Deal'])
                
                st.info(f"Processed watchlist data with {len(watchlist_df)} rows")
                
                # Perform left join with main dataframe
                result_df = main_df.merge(
                    watchlist_df[['Deal', 'Watchlist']], 
                    left_on='Investment name', 
                    right_on='Deal', 
                    how='left'
                )
                
                # Drop the redundant Deal column
                if 'Deal' in result_df.columns:
                    result_df = result_df.drop(columns=['Deal'])
                
                st.success("Successfully joined watchlist data")
                return result_df
            else:
                st.warning("Unexpected column count in watchlist file")
                return main_df
        else:
            st.warning("Not enough columns in watchlist file")
            return main_df
            
    except Exception as e:
        st.error(f"Error processing watchlist file: {e}")
        st.warning("Returning main dataframe without watchlist data")
        return main_df

def read_delta_master_eu(file_bytes, asset_manager, year, reporting_date, quarter):
    """
    Read and process the EU Delta Master file with additional calculated columns.
    """
    # List of original columns to extract
    original_columns = [
        'Asset Subportfolio',
        'Repayment year',
        'Investment style',
        '32_PROPERTY COUNTRY',
        '14 Floating Fixed rate',
        'Exposure cum',
        'Current loan amount cum',
        'Open commitment cum',
        'Main Sector',
        'Remaining life',
        'ARE CREL Grade',
        'Interest rate / Coupon p.a. (%)',
        'LTV (%)',
        'DSCR',
        'ICR',
        'Net Debt Yield (%)'
    ]
    
    # New column names to rename to
    new_columns = [
        'Investment name',
        'Repayment',
        'Investment style',
        'Country',
        'Fixed vs. Floating',
        'Loan commitment EUR',
        'Loan amount drawn EUR',
        'Loan amount undrawn/ repaid EUR',
        'Sector',
        'Remaining loan term if fully extended',
        'Credit Rating',
        'Current coupon (p.a.)',
        'Latest LTV',
        'DSCR (current) T12 capped',
        'ICR (current) T12 capped',
        'Debt yield (current)'
    ]
    
    # Create mapping from original column names to new column names
    column_mapping = dict(zip(original_columns, new_columns))
    
    try:
        # Load the first sheet without specifying a header
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=0, header=None)
        
        # Helper function to count consecutive NaNs at the start of a row
        def count_consecutive_nans_at_start(row):
            count = 0
            for val in row:
                if pd.isna(val):
                    count += 1
                else:
                    break
            return count
        
        # Iterate through rows to find the header row
        header_found = False
        for i, row in df.iterrows():
            consecutive_nans = count_consecutive_nans_at_start(row)
            valid_headers = sum(1 for cell in row if not pd.isna(cell) and not str(cell).startswith("Unnamed"))
            
            # Check conditions for a valid header row
            if (valid_headers >= 3 and consecutive_nans <= 5):
                st.info(f"Header row detected at index {i}")
                header_found = True
                
                # Read the file again with the detected header row
                df_with_headers = pd.read_excel(BytesIO(file_bytes), sheet_name=0, header=i)
                
                # Clean the headers by replacing '\n' with ' '
                df_with_headers.columns = [col.replace('\n', ' ') if isinstance(col, str) else col for col in df_with_headers.columns]
                
                # Forward fill NaN values in text/object columns and columns with 'year' in the header
                for col in df_with_headers.columns:
                    if (df_with_headers[col].dtype == 'object') or (
                            isinstance(col, str) and 'year' in col.lower()):
                        # Use ffill() instead of fillna(method='ffill') to avoid deprecation warning
                        df_with_headers[col] = df_with_headers[col].ffill()
                
                # Check which required columns exist in the dataframe
                available_columns = [col for col in original_columns if col in df_with_headers.columns]
                missing_columns = [col for col in original_columns if col not in df_with_headers.columns]
                
                if missing_columns:
                    st.warning(f"The following requested columns were not found in the file: {missing_columns}")
                
                if not available_columns:
                    st.error("None of the requested columns were found in the file.")
                    return pd.DataFrame()
                
                # Select only the required columns that exist in the dataframe
                result_df = df_with_headers[available_columns].copy()
                
                # Remove rows where 'Asset Subportfolio' equals 'Sum'
                if 'Asset Subportfolio' in available_columns:
                    initial_row_count = len(result_df)
                    result_df = result_df[result_df['Asset Subportfolio'] != 'Sum']
                    rows_removed = initial_row_count - len(result_df)
                    st.info(f"Removed {rows_removed} rows where 'Asset Subportfolio' = 'Sum'")
                  # Rename the columns according to the mapping
                rename_mapping = {old_col: column_mapping[old_col] for old_col in available_columns}
                result_df = result_df.rename(columns=rename_mapping)
                
                # Apply property_type_mapping to the Sector column
                property_type_mapping = {
                    'Mixed': 'Other',
                    'MIXED': 'Other',
                    'Various': 'Other',
                    'MEDICAL': 'Other',
                    'MULTIFAM': 'Residential',
                    'WAREHOUSE': 'Logistics',
                    'RETAIL': 'Retail',
                    'OFFICE': 'Office',
                    'Storage': 'Other',
                    'STORAGE': 'Other',
                    'APARTMENT': 'Residential',
                    'Logistic': 'Logistics'
                }
                
                if 'Sector' in result_df.columns:
                    result_df['Sector'] = result_df['Sector'].apply(
                        lambda x: property_type_mapping.get(x, x) if pd.notna(x) else x
                    )
                
                # Convert numeric columns to appropriate types for calculations
                numeric_columns = [
                    'Loan commitment EUR', 'Loan amount drawn EUR', 'Latest LTV', 
                    'Current coupon (p.a.)', 'DSCR (current) T12 capped', 
                    'ICR (current) T12 capped', 'Debt yield (current)'
                ]
                
                for col in numeric_columns:
                    if col in result_df.columns:
                        # Convert percentage columns to decimals for calculations
                        if 'LTV' in col or 'coupon' in col or 'yield' in col:
                            result_df[col] = pd.to_numeric(result_df[col].astype(str).str.replace('%', '').str.replace(',', '.'), errors='coerce') / 100
                        else:
                            result_df[col] = pd.to_numeric(result_df[col].astype(str).str.replace(',', '.'), errors='coerce')                
                # Multiply by 100 to convert decimal to percentage
                result_df['Latest LTV'] = result_df['Latest LTV']*100
                result_df['Current coupon (p.a.)'] = result_df['Current coupon (p.a.)']*100

                # Add new calculated columns
                # 1. Year
                result_df['Year'] = year
                
                # 2. Reporting Date
                result_df['Reporting Date'] = reporting_date
                
                # 3. Quarter
                result_df['Quarter'] = quarter
                
                # 4. Asset manager (capitalized)
                result_df['Asset manager'] = asset_manager.upper()
                
                # 5. Strategy: based on Investment style
                def determine_strategy(style):
                    if pd.isna(style):
                        return None
                    style_lower = str(style).lower()
                    if 'core' in style_lower:
                        return 'PPRE EU Core'
                    elif 'enhanced' in style_lower:
                        return 'PPRE EU Enhanced'
                    return None
                
                result_df['Strategy'] = result_df['Investment style'].apply(determine_strategy)
                
                # 6. Development: based on Investment style
                def determine_development(style):
                    if pd.isna(style):
                        return None
                    style_lower = str(style).lower()
                    if 'core' in style_lower:
                        return 'Stabilized'
                    elif 'enhanced' in style_lower:
                        return 'Development'
                    return None
                
                result_df['Development'] = result_df['Investment style'].apply(determine_development)
                
                # 7. Seniority: if asset_manager.upper() == 'PPRE EU' then 'Senior'
                if asset_manager.upper() == 'PPRE EU':
                    result_df['Seniority'] = 'Senior'
                else:
                    result_df['Seniority'] = None                # 8. Current coupon (p.a.) weighted
                if 'Current coupon (p.a.)' in result_df.columns and 'Loan amount drawn EUR' in result_df.columns:
                    result_df['Current coupon (p.a.) weighted'] = result_df['Current coupon (p.a.)'] * result_df['Loan amount drawn EUR']
                  # 9. Latest LTV Weighted
                if 'Latest LTV' in result_df.columns and 'Loan amount drawn EUR' in result_df.columns:
                    result_df['Latest LTV Weighted'] = result_df['Loan amount drawn EUR'] * result_df['Latest LTV']
                  # 10-13. LTV range columns
                if 'Latest LTV' in result_df.columns:
                    result_df['LTV < 55%'] = result_df['Latest LTV'].apply(lambda x: x if pd.notna(x) and x < 0.55 else None)
                    result_df['55% ≤ LTV < 65%'] = result_df['Latest LTV'].apply(lambda x: x if pd.notna(x) and 0.55 <= x < 0.65 else None)
                    result_df['65% ≤ LTV < 75%'] = result_df['Latest LTV'].apply(lambda x: x if pd.notna(x) and 0.65 <= x < 0.75 else None)
                    result_df['75% < LTV'] = result_df['Latest LTV'].apply(lambda x: x if pd.notna(x) and x >= 0.75 else None)
                
                # 14. DSCR (current) Weighted
                if 'DSCR (current) T12 capped' in result_df.columns and 'Loan amount drawn EUR' in result_df.columns:
                    result_df['DSCR (current) Weighted'] = result_df['Loan amount drawn EUR'] * result_df['DSCR (current) T12 capped']
                
                # 15. ICR (current) Weighted
                if 'ICR (current) T12 capped' in result_df.columns and 'Loan commitment EUR' in result_df.columns:
                    result_df['ICR (current) Weighted'] = result_df['Loan commitment EUR'] * result_df['ICR (current) T12 capped']
                  # 16. Debt yield (current) Weighted
                if 'Debt yield (current)' in result_df.columns and 'Loan amount drawn EUR' in result_df.columns:
                    result_df['Debt yield (current)'] = result_df['Debt yield (current)'] * 100
                    result_df['Debt yield (current) Weighted'] = result_df['Loan amount drawn EUR'] * result_df['Debt yield (current)']
                
                # 17. Impairment
                if asset_manager.upper() == 'PPRE EU':
                    result_df['Impairment'] = 'No'
                else:
                    result_df['Impairment'] = None
                
                # 18. Correct value in 'Country' column
                if asset_manager.upper() == 'PPRE EU' and 'Country' in result_df.columns:
                    result_df['Country'] = result_df['Country'].apply(lambda x: 'UK' if x == 'GB' else x)
                
                # 19. "Loan CCY" column based on Country 
                if asset_manager.upper() == 'PPRE EU' and 'Country' in result_df.columns:
                    result_df['Loan CCY'] = result_df['Country'].apply(lambda x: 'GBP' if x == 'UK' else 'EUR')
                
                # 20. Add empty columns for PPRE EU
                if asset_manager.upper() == 'PPRE EU':
                    empty_columns = [
                        "Identifier BNY Report 'Issuer Name'",
                        "Loan commitment (loan CCY)",
                        "Loan amount drawn (loan CCY)",
                        "Loan amount undrawn/ repaid (loan CCY)",
                        "Remaining loan term w/o extension",
                        "Loan valuation"
                    ]
                    
                    for col in empty_columns:
                        result_df[col] = None
                
                st.success("Successfully added all calculated columns")
                return result_df
        
        if not header_found:
            st.error("No valid header row found in the file.")
            return pd.DataFrame()
            
    except Exception as e:
        st.error(f"Error processing the file: {e}")
        return pd.DataFrame()

def format_excel_worksheet(ws):
    """
    Apply beautiful formatting to Excel worksheet without using Table objects
    """
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    even_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
    
    # Get dimension holder
    dim_holder = DimensionHolder(worksheet=ws)
    
    # Format header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Set column width based on header content (with padding)
        if cell.value:
            col_letter = cell.column_letter
            # Calculate width based on header length with padding
            header_width = max(10, min(30, len(str(cell.value)) + 4))
            dim_holder[col_letter] = ColumnDimension(ws, index=col_letter, width=header_width)
    
    # Apply dimension holder to worksheet
    ws.column_dimensions = dim_holder
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Format data rows and apply number formats
    currency_columns = [
        'Loan commitment (loan CCY)', 'Loan amount drawn (loan CCY)', 
        'Loan amount undrawn/ repaid (loan CCY)', 'Loan commitment EUR', 
        'Loan amount drawn EUR', 'Loan amount undrawn/ repaid EUR'
    ]
    
    percent_columns = [
        'Current coupon (p.a.)', 'Latest LTV', 'Debt yield (current)',
        'LTV < 55%', '55% ≤ LTV < 65%', '65% ≤ LTV < 75%', '75% < LTV',
        'Debt yield (current) Weighted', 'Latest LTV Weighted', 'Current coupon (p.a.) weighted'
    ]
    
    decimal_columns = [
        'DSCR (current) T12 capped', 'ICR (current) T12 capped',
        'Remaining loan term w/o extension', 'Remaining loan term if fully extended'
    ]
    
    # Find column indexes
    headers = [cell.value for cell in ws[1]]
    
    # Format each row with alternating colors
    for row in range(2, ws.max_row + 1):
        # Apply alternating row colors
        row_fill = even_fill if row % 2 == 0 else None
        
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=i)
            cell.border = thin_border
            
            if row_fill:
                cell.fill = row_fill
            
            # Apply specific formatting based on column type
            if header in currency_columns and cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
            elif header in percent_columns and cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00%'
            elif header in decimal_columns and cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00'
    
    # Add auto-filter but not as a Table
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    
    return ws

def append_to_template(df, template_file_bytes):
    """
    Simplified version that appends the processed data to template 
    without advanced formatting that might cause corruption.
    """
    try:
        # Read template into pandas
        with BytesIO(template_file_bytes) as template_buffer:
            # Get all sheet names
            xls = pd.ExcelFile(template_buffer)
            sheet_names = xls.sheet_names
            
            if 'Data_hub' not in sheet_names:
                st.error("The template file does not contain a 'Data_hub' sheet.")
                return None
            
            # Read all sheets into a dict of dataframes
            sheet_dfs = {}
            for sheet in sheet_names:
                # Reset buffer position
                template_buffer.seek(0)
                sheet_dfs[sheet] = pd.read_excel(template_buffer, sheet_name=sheet)
            
            # Get the Data_hub sheet
            data_hub_df = sheet_dfs['Data_hub']
            
            # Define key columns for duplicate checking
            key_columns = ['Investment name', 'Asset manager', 'Reporting Date']
            existing_keys = set()
            
            # Create keys from existing data
            for _, row in data_hub_df.iterrows():
                key_tuple = tuple(str(row.get(col, '')) for col in key_columns if col in data_hub_df.columns)
                if any(key_tuple):
                    existing_keys.add(key_tuple)
            
            # Filter out rows from new data that already exist
            new_rows = []
            for _, row in df.iterrows():
                key_tuple = tuple(str(row.get(col, '')) for col in key_columns if col in df.columns)
                if not (key_tuple in existing_keys and any(key_tuple)):
                    new_rows.append(row)
            
            if not new_rows:
                st.info("No new rows added. All data already exists in the template.")
                # Return a copy of the original template
                output = BytesIO()
                output.write(template_file_bytes)
                output.seek(0)
                return output
            
            # Create DataFrame from new rows
            new_df = pd.DataFrame(new_rows)
            
            # Make sure both dataframes have the same columns to avoid warnings
            all_columns = list(set(data_hub_df.columns) | set(new_df.columns))
            for col in all_columns:
                if col not in data_hub_df.columns:
                    data_hub_df[col] = None
                if col not in new_df.columns:
                    new_df[col] = None
            
            # Append new data to Data_hub df
            updated_data_hub = pd.concat([data_hub_df, new_df], axis=0, ignore_index=True)
            sheet_dfs['Data_hub'] = updated_data_hub
            
            # Write all sheets to a new Excel file
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                for sheet_name, sheet_df in sheet_dfs.items():
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            st.info(f"Added {len(new_rows)} new rows to the Data_hub sheet")
            output.seek(0)
            return output
            
    except Exception as e:
        st.error(f"Error appending to template: {e}")
        import traceback
        st.error(traceback.format_exc())
        return None

def append_to_template_fast(df, template_file_bytes):
    """
    Faster version that appends data to template directly using openpyxl
    without reading all sheets into pandas dataframes.
    """
    try:
        # Load the workbook directly with openpyxl
        with BytesIO(template_file_bytes) as template_buffer:
            wb = openpyxl.load_workbook(template_buffer)
            
            # Check if Data_hub exists
            if 'Data_hub' not in wb.sheetnames:
                st.error("The template file does not contain a 'Data_hub' sheet.")
                return None
            
            # Get the Data_hub sheet
            ws = wb['Data_hub']
            
            # Get headers from the first row
            headers = [cell.value for cell in ws[1]]
            
            # Define key columns for duplicate checking
            key_columns = ['Investment name', 'Asset manager', 'Reporting Date']
            key_indices = [headers.index(key) + 1 for key in key_columns if key in headers]
            
            # Read existing keys
            existing_keys = set()
            max_row = ws.max_row
            
            # Find the true max row (excluding empty trailing rows)
            while max_row > 1:
                if any(ws.cell(row=max_row, column=i).value for i in range(1, len(headers) + 1)):
                    break
                max_row -= 1
            
            # Collect existing keys
            for row in range(2, max_row + 1):
                key_values = tuple(str(ws.cell(row=row, column=i).value or '') for i in key_indices)
                if any(key_values):
                    existing_keys.add(key_values)
            
            # Filter rows to append
            rows_to_append = []
            for _, row_data in df.iterrows():
                key_values = tuple(str(row_data.get(headers[i-1], '') or '') for i in key_indices)
                if not (key_values in existing_keys and any(key_values)):
                    rows_to_append.append(row_data)
            
            if not rows_to_append:
                st.info("No new rows added. All data already exists in the template.")
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                return output
                
            # Add new rows directly to worksheet
            for row_data in rows_to_append:
                max_row += 1
                for col_idx, header in enumerate(headers, 1):
                    if header in df.columns:
                        value = row_data.get(header)
                        ws.cell(row=max_row, column=col_idx, value=value)
            
            # Save the workbook to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            st.info(f"Added {len(rows_to_append)} new rows to the Data_hub sheet")
            return output
            
    except Exception as e:
        st.error(f"Error in fast template append: {e}")
        import traceback
        st.error(traceback.format_exc())
        return None

def beautify_excel_output(template_bytes):
    """
    Apply beautiful formatting to the output template without using Table objects
    that can cause corruption.
    """
    try:
        # Read the template into a workbook
        wb = openpyxl.load_workbook(BytesIO(template_bytes))
        
        # Check if Data_hub exists
        if 'Data_hub' not in wb.sheetnames:
            return BytesIO(template_bytes)
        
        # Get the Data_hub sheet
        ws = wb['Data_hub']
        
        # Apply formatting if there's data
        if ws.max_row > 1:
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            even_fill = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
            
            # Format header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Format data rows with alternating colors (no Table object)
            for row_idx in range(2, ws.max_row + 1):
                # Apply alternating row colors
                if row_idx % 2 == 0:  # Even rows
                    for cell in ws[row_idx]:
                        cell.fill = even_fill
                        cell.border = thin_border
                else:  # Odd rows
                    for cell in ws[row_idx]:
                        cell.border = thin_border
            
            # Add auto-filter without creating a Table
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            
            # Freeze panes
            ws.freeze_panes = 'A2'
            
            # Auto-fit columns and apply number formatting
            currency_columns = ['Loan commitment (loan CCY)', 'Loan amount drawn (loan CCY)', 
                               'Loan amount undrawn/ repaid (loan CCY)', 'Loan commitment EUR', 
                               'Loan amount drawn EUR', 'Loan amount undrawn/ repaid EUR']
            
            percent_columns = ['Current coupon (p.a.)', 'Latest LTV', 'Debt yield (current)',
                           'Debt yield (current) Weighted', 'Latest LTV Weighted', 'Current coupon (p.a.) weighted']
            
            # Get column headers
            headers = [cell.value for cell in ws[1]]
            
            for column in ws.columns:
                # Auto-fit width
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                column_idx = column[0].column - 1  # 0-based index
                
                for cell in column:
                    try:
                        if len(str(cell.value or '')) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = max(max_length + 2, 10)
                ws.column_dimensions[column_letter].width = min(adjusted_width, 30)
                
                # Apply number formatting for specific columns
                if column_idx < len(headers):
                    header = headers[column_idx]
                    if header in currency_columns:
                        for cell in column[1:]:  # Skip header
                            if isinstance(cell.value, (int, float)) and cell.value is not None:
                                cell.number_format = '#,##0.00'
                    elif header in percent_columns:
                        for cell in column[1:]:  # Skip header
                            if isinstance(cell.value, (int, float)) and cell.value is not None:
                                cell.number_format = '0.00%'
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
        
    except Exception as e:
        # If any error occurs, return the original template
        st.warning(f"Error beautifying template: {e}. Using basic format.")
        return BytesIO(template_bytes)

def create_excel_output(df, output_filename="processed_data.xlsx"):
    """
    Create a beautifully formatted Excel file from the DataFrame without using Table objects
    """
    try:
        # Create a BytesIO object for the Excel file
        output_buffer = BytesIO()
        
        # Save the DataFrame to an Excel writer
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data')
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = workbook['Data']
            
            # Apply formatting to the worksheet using our modified function
            format_excel_worksheet(worksheet)
            
        # Reset buffer position
        output_buffer.seek(0)
        return output_buffer
        
    except Exception as e:
        st.error(f"Error creating Excel output: {e}")
        import traceback
        st.error(traceback.format_exc())
        return None

def create_download_zip(result_df, modified_template=None):
    """
    Create a zip file containing all output files
    
    Args:
        result_df: Processed DataFrame
        modified_template: Modified template file (optional)
        
    Returns:
        BytesIO: Zip file as bytes
    """
    try:
        # Create a BytesIO object to store the zip file
        zip_buffer = BytesIO()
        
        # Create the formatted Excel output
        excel_data = create_excel_output(result_df)
        
        # Create a zip file
        with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED, False) as zip_file:
            # Add the processed data as CSV
            csv_data = result_df.to_csv(index=False)
            zip_file.writestr("processed_data.csv", csv_data)
            
            # Add the processed data as Excel (beautifully formatted)
            if excel_data:
                zip_file.writestr("processed_data.xlsx", excel_data.getvalue())
            
            # Add the modified template if available
            if modified_template is not None:
                zip_file.writestr("updated_template.xlsx", modified_template.getvalue())
        
        # Reset buffer position
        zip_buffer.seek(0)
        return zip_buffer
    
    except Exception as e:
        st.error(f"Error creating zip file: {e}")
        return None

def process_combined_files(reporting_date, all_files, watchlist_file, template_file):
    """
    Process both PPRE EU and PPRE US files in a single run.
    
    Args:
        reporting_date: Reporting date string (DD.MM.YYYY)
        all_files: List of uploaded files
        watchlist_file: Watchlist file (for PPRE EU data)
        template_file: Template file for output
        
    Returns:
        tuple: (final_df, modified_template)
    """
    # Parse reporting date
    date_obj = datetime.strptime(reporting_date, "%d.%m.%Y")
    year = date_obj.year
    month = date_obj.month
    
    # Determine quarter
    if 1 <= month <= 3:
        quarter = "Q1"
    elif 4 <= month <= 6:
        quarter = "Q2"
    elif 7 <= month <= 9:
        quarter = "Q3"
    else:
        quarter = "Q4"
    
    # Determine sheet name for watchlist file (e.g., "2024Q4")
    watchlist_sheet = f"{year}{quarter}"
    
    # Get exchange rate for PPRE US case
    usd_eur_fx = get_exchange_rate(date_obj, method="ecb")
    st.info(f"Using USD/EUR exchange rate: {usd_eur_fx} for {reporting_date}")
    
    # Separate files into EU and US
    delta_files = [f for f in all_files if 'Delta' in f.name]
    azl_files = [f for f in all_files if 'AZL' in f.name]
    
    # Process remaining files based on name patterns
    remaining_files = [f for f in all_files if 'Delta' not in f.name and 'AZL' not in f.name]
    
    # Process EU files
    st.subheader("Processing PPRE EU Files")
    eu_results = []
    
    if delta_files:
        for delta_file in delta_files:
            st.write(f"Processing EU file: {delta_file.name}...")
            result_df = read_delta_master_eu(delta_file.getvalue(), "PPRE EU", year, reporting_date, quarter)
            if not result_df.empty and watchlist_file is not None:
                result_df = integrate_watchlist(result_df, watchlist_file.getvalue(), watchlist_sheet)
            if not result_df.empty:
                eu_results.append(result_df)
    else:
        st.warning("No files with 'Delta' in the name were found for PPRE EU processing.")
    
    # Process US files
    st.subheader("Processing PPRE US Files")
    us_results = []
    
    if azl_files:
        for azl_file in azl_files:
            st.write(f"Processing US file: {azl_file.name}...")
            result_df = read_delta_master_us(azl_file.getvalue(), "PPRE US", year, reporting_date, quarter, usd_eur_fx, date_obj)
            if not result_df.empty:
                us_results.append(result_df)
    else:
        st.warning("No files with 'AZL' in the name were found for PPRE US processing.")
    
    # Process any remaining files by trying both processors
    if remaining_files:
        st.subheader("Processing Other Files")
        for file in remaining_files:
            st.write(f"Attempting to process: {file.name}...")
            
            # Try EU processor first
            eu_result = read_delta_master_eu(file.getvalue(), "PPRE EU", year, reporting_date, quarter)
            if not eu_result.empty:
                st.write(f"Successfully processed as PPRE EU file")
                if watchlist_file is not None:
                    eu_result = integrate_watchlist(eu_result, watchlist_file.getvalue(), watchlist_sheet)
                eu_results.append(eu_result)
                continue
                
            # Try US processor if EU fails
            us_result = read_delta_master_us(file.getvalue(), "PPRE US", year, reporting_date, quarter, usd_eur_fx, date_obj)
            if not us_result.empty:
                st.write(f"Successfully processed as PPRE US file")
                us_results.append(us_result)
                continue
                
            st.warning(f"Could not process {file.name} with either EU or US processor")
    
    # Combine all results
    all_results = eu_results + us_results
    
    if not all_results:
        st.error("No data was successfully processed from any of the uploaded files.")
        return None, None
    
    # Use safe concat to avoid warning
    final_df = safe_concat(all_results)
    
    # Final column selection and ordering
    final_columns = [
        'Reporting Date',
        'Year',
        'Quarter',
        'Asset manager',
        'Strategy',
        "Identifier BNY Report 'Issuer Name'",
        'Investment name',
        'Loan CCY',
        'Loan commitment (loan CCY)',
        'Loan amount drawn (loan CCY)',
        'Loan amount undrawn/ repaid (loan CCY)',
        'Loan commitment EUR',
        'Loan amount drawn EUR',
        'Loan amount undrawn/ repaid EUR',
        'Development',
        'Country',
        'Sector',
        'Fixed vs. Floating',
        'Seniority',
        'Remaining loan term w/o extension',
        'Remaining loan term if fully extended',
        'Credit Rating',
        'Loan valuation',
        'Current coupon (p.a.)',
        'Current coupon (p.a.) weighted',
        'Latest LTV',
        'Latest LTV Weighted',
        'LTV < 55%',
        '55% ≤ LTV < 65%',
        '65% ≤ LTV < 75%',
        '75% < LTV',
        'DSCR (current) T12 capped',
        'DSCR (current) Weighted',
        'ICR (current) T12 capped',
        'ICR (current) Weighted',
        'Debt yield (current)',
        'Debt yield (current) Weighted',
        'Watchlist',
        'Impairment'
    ]
    
    # Check which columns exist in the final_df
    available_columns = [col for col in final_columns if col in final_df.columns]
    missing_columns = [col for col in final_columns if col not in final_df.columns]
    
    if missing_columns:
        st.warning(f"The following columns were not available and will be created as empty: {missing_columns}")
        for col in missing_columns:
            final_df[col] = None
    
    # Select only the required columns in the specified order
    final_result = final_df[final_columns].copy()

    # Sort by Asset manager
    #final_result = final_result.sort_values(by='Asset manager')
    
    st.success(f"Final combined DataFrame contains {len(final_result)} rows and {len(final_columns)} columns")
    
    # If template file provided, append data to it
    if template_file is not None:
        basic_template = append_to_template_fast(final_result, template_file.getvalue())
        if basic_template:
            modified_template = beautify_excel_output(basic_template.getvalue())
        else:
            modified_template = None
        return final_result, modified_template
    else:
        return final_result, None

def process_files(asset_manager, reporting_date, input_files, watchlist_file=None, template_file=None):
    """
    Process uploaded files and return the final dataframe
    """
    try:
        # Parse reporting date
        date_obj = datetime.strptime(reporting_date, "%d.%m.%Y")
        year = date_obj.year
        month = date_obj.month
        
        # Determine quarter
        if 1 <= month <= 3:
            quarter = "Q1"
        elif 4 <= month <= 6:
            quarter = "Q2"
        elif 7 <= month <= 9:
            quarter = "Q3"
        else:
            quarter = "Q4"
        
        # Format reporting date as mm.dd.yyyy
        formatted_reporting_date = date_obj.strftime("%m.%d.%Y")
        
        # Determine sheet name for watchlist file (e.g., "2024Q4")
        watchlist_sheet = f"{year}{quarter}"
        
        # For combined processing, use the dedicated function
        if asset_manager == "PPRE EU + PPRE US":
            return process_combined_files(reporting_date, input_files, watchlist_file, template_file)
        
        # Get exchange rate for reporting date (for PPRE US case)
        usd_eur_fx = get_exchange_rate(date_obj, method="ecb")
        st.info(f"Using USD/EUR exchange rate: {usd_eur_fx} for {reporting_date}")
        
        # Process based on asset manager
        if asset_manager == "PPRE EU":
            if not input_files:
                st.error("Please upload at least one Delta Master file.")
                return None, None
                
            # Find files with 'Delta' in the name
            delta_files = [f for f in input_files if 'Delta' in f.name]
            
            if not delta_files:
                st.warning("No files with 'Delta' in the name were found. Processing all uploaded files as Delta Master files.")
                delta_files = input_files
                
            # Process all delta files and combine
            all_results = []
            for delta_file in delta_files:
                st.write(f"Processing {delta_file.name}...")
                result_df = read_delta_master_eu(delta_file.getvalue(), asset_manager, year, reporting_date, quarter)
                
                if not result_df.empty and watchlist_file is not None:
                    # Process watchlist file and join with main dataframe
                    result_df = integrate_watchlist(result_df, watchlist_file.getvalue(), watchlist_sheet)
                
                if not result_df.empty:
                    all_results.append(result_df)
            
            if not all_results:
                st.error("No data was successfully processed from the uploaded files.")
                return None, None
                
            # Use safe concat to avoid warning
            final_df = safe_concat(all_results)
                
        elif asset_manager == "PPRE US":
            if not input_files:
                st.error("Please upload at least one AZL file.")
                return None, None
                
            # Find files with 'AZL' in the name
            azl_files = [f for f in input_files if 'AZL' in f.name]
            
            if not azl_files:
                st.warning("No files with 'AZL' in the name were found. Processing all uploaded files as US files.")
                azl_files = input_files
                
            # Process all AZL files and combine
            all_results = []
            for azl_file in azl_files:
                st.write(f"Processing {azl_file.name}...")
                result_df = read_delta_master_us(azl_file.getvalue(), asset_manager, year, 
                                                reporting_date, quarter, usd_eur_fx, date_obj)
                
                if not result_df.empty:
                    all_results.append(result_df)
            
            if not all_results:
                st.error("No data was successfully processed from the uploaded files.")
                return None, None
                
            # Use safe concat to avoid warning
            final_df = safe_concat(all_results)
            
        else:
            st.error(f"Unsupported asset manager '{asset_manager}'. Only 'PPRE EU', 'PPRE US', and 'PPRE EU + PPRE US' are supported.")
            return None, None
        
        # Final column selection and ordering
        final_columns = [
            'Reporting Date',
            'Year',
            'Quarter',
            'Asset manager',
            'Strategy',
            "Identifier BNY Report 'Issuer Name'",
            'Investment name',
            'Loan CCY',
            'Loan commitment (loan CCY)',
            'Loan amount drawn (loan CCY)',
            'Loan amount undrawn/ repaid (loan CCY)',
            'Loan commitment EUR',
            'Loan amount drawn EUR',
            'Loan amount undrawn/ repaid EUR',
            'Development',
            'Country',
            'Sector',
            'Fixed vs. Floating',
            'Seniority',
            'Remaining loan term w/o extension',
            'Remaining loan term if fully extended',
            'Credit Rating',
            'Loan valuation',
            'Current coupon (p.a.)',
            'Current coupon (p.a.) weighted',
            'Latest LTV',
            'Latest LTV Weighted',
            'LTV < 55%',
            '55% ≤ LTV < 65%',
            '65% ≤ LTV < 75%',
            '75% < LTV',
            'DSCR (current) T12 capped',
            'DSCR (current) Weighted',
            'ICR (current) T12 capped',
            'ICR (current) Weighted',
            'Debt yield (current)',
            'Debt yield (current) Weighted',
            'Watchlist',
            'Impairment'
        ]
        
        # Check which columns exist in the final_df
        available_columns = [col for col in final_columns if col in final_df.columns]
        missing_columns = [col for col in final_columns if col not in final_df.columns]
        
        if missing_columns:
            st.warning(f"The following columns were not available and will be created as empty: {missing_columns}")
            for col in missing_columns:
                final_df[col] = None
        
        # Select only the required columns in the specified order
        final_result = final_df[final_columns].copy()

        # Sort by Asset manager
        # final_result = final_result.sort_values(by='Asset manager')
        
        st.success(f"Final DataFrame contains {len(final_result)} rows and {len(final_columns)} columns")
        
        # If template file provided, append data to it
        if template_file is not None:
            basic_template = append_to_template(final_result, template_file.getvalue())
            if basic_template:
                modified_template = beautify_excel_output(basic_template.getvalue())
            else:
                modified_template = None
            return final_result, modified_template
        else:
            return final_result, None
        
    except Exception as e:
        st.error(f"Error processing files: {e}")
        import traceback
        st.error(traceback.format_exc())
        return None, None

def show_enhanced_documentation():
    """
    Display detailed documentation about the data processing with enhanced transformation details
    """
    st.header("Delta Master Processor Documentation")
    
    st.subheader("Overview")
    st.write("""
    This application processes financial data from asset managers (PPRE EU and PPRE US) 
    and transforms it into a standardized format for analysis and reporting.
    """)
    
    tabs = st.tabs(["PPRE EU Process", "PPRE US Process", "Combined Processing", "Column Mappings", "Data Transformations", "Calculations", "Technical Details"])
    
    with tabs[0]:
        st.markdown("""
        ### PPRE EU Processing Flow
        
        **Input Files Required:**
        - **Delta Master File** - Contains loan data. Filename should contain 'Delta'
        - **Watchlist File** - Contains watchlist/traffic light data
        
        **Processing Steps:**
        1. **Header Detection**
           - Automatically finds the header row in the Delta Master file
           - Headers are detected when there are at least 3 valid headers with no more than 5 consecutive NaN values at the beginning
           - This provides flexibility to handle files with various header positions
        
        2. **Data Cleaning**
           - Forward fills text columns and columns with 'year' in the name to handle hierarchical data
           - Removes rows where 'Asset Subportfolio' = 'Sum' to avoid duplicate counting
           - Converts percentage values to decimals (divides by 100) for proper calculations
           - Handles regional formatting issues (commas, dots in numbers)
        
        3. **Column Renaming**
           - Maps original column names to standardized names for consistency across regions
           - Maintains naming conventions for cross-reporting compatibility
        
        4. **Calculated Columns**
           - Adds reporting period information (Year, Quarter, Reporting Date) for time-based analysis
           - Determines Strategy from Investment style (PPRE EU Core, PPRE EU Enhanced)
           - Determines Development status from Investment style (Stabilized, Development)
           - Sets Seniority to 'Senior' for PPRE EU investments
           - Calculates weighted metrics using loan amounts as weights
           - Generates LTV range columns for risk classification
        
        5. **Country Correction**
           - Changes 'GB' to 'UK' for consistency across systems
           - Sets 'Loan CCY' to 'UKP' for UK investments and 'EUR' for others
        
        6. **Watchlist Integration**
           - Joins with watchlist data on 'Investment name' = 'Deal'
           - Sets 'Watchlist' based on 'Traffic light' values: Green -> No, Red/Yellow -> Yes
           - Ensures risk flag consistency
        """)
        
    with tabs[1]:
        st.markdown("""
        ### PPRE US Processing Flow
        
        **Input Files Required:**
        - **AZL File** - Contains US loan data. Filename should contain 'AZL'
        
        **Processing Steps:**
        1. **Data Filtering**
           - Selects rows where 'Asset Manager' is 'ARE' or 'JPM'
           - Changes 'ARE' to 'PPRE US' for consistent naming
           - Filters out irrelevant data early in the process
        
        2. **Data Standardization**
           - Cleans 'Watchlist' values: Red/Yellow -> Yes, others -> No
           - Normalizes LTV values by dividing by 100 if not already in decimal format
           - Maps property types to standardized categories for consistent sector analysis:
             * Mixed, Various, MIXED -> Other
             * MULTIFAM -> Residential
             * WAREHOUSE -> Logistics
             * RETAIL -> Retail
             * OFFICE -> Office
             * Storage -> Other
             * STORAGE -> Other
             * APARTMENT -> Residential
             * Logistic -> Logistics
        
        3. **Column Renaming**
           - Maps US-specific column names to standardized names
           - Ensures compatibility with EU reporting formats
        
        4. **Data Transformation**
           - Converts USD values to EUR using the retrieved exchange rate
           - Calculates loan amounts (committed, drawn, undrawn) in both currencies
           - Determines remaining loan terms in years from maturity date with exact day calculation
           - Sets 'Country' to 'US' and 'Seniority' to 'Senior'
           - Calculates weighted metrics based on loan amounts
           - Categorizes investments with Development flag based on 'CONSTRUCTION' value
        
        5. **Financial Calculations**
           - Calculates Interest Coverage Ratio (ICR) from Net Operating Income and interest payments
           - Calculates Debt Yield from NOI and loan amount
           - Creates multiple LTV category columns for risk assessment
           - Computes weighted values for key metrics to enable portfolio-level analysis
        """)
    
    with tabs[2]:
        st.markdown("""
        ### Combined Processing Mode
        
        The "PPRE EU + PPRE US" option processes all uploaded files in a single run:
        
        **Features:**
        - Automatically categorizes files with "Delta" in the name as PPRE EU files
        - Automatically categorizes files with "AZL" in the name as PPRE US files
        - Attempts to process any other files with both processors to determine the correct format
        - Applies the appropriate processor to each file type
        - Combines all results into a single DataFrame with consistent structure
        - Applies the watchlist data to PPRE EU files
        - Creates a consistent output format combining both data sources
        
        **Process Flow:**
        1. **File Categorization**
           - Automatically sorts uploaded files by name pattern
           - Routes each file to the correct processor
        
        2. **Parallel Processing**
           - Processes EU and US files with their respective logic
           - Maintains region-specific calculations and transformations
        
        3. **Data Integration**
           - Safely concatenates results with consistent column structure
           - Applies column standardization across different data sources
           - Ensures all data has the same type and format before combination
        
        4. **Output Generation**
           - Creates a unified output with data from all sources
           - Maintains consistent formatting and structure
           - Preserves region-specific information while providing global view
        """)
    
    with tabs[3]:
        st.markdown("""
        ### Column Mappings
        
        **PPRE EU Column Mapping:**
        
        | Original Column | Standardized Column | Description |
        |----------------|---------------------|-------------|
        | Asset Subportfolio | Investment name | Primary identifier of the loan |
        | Repayment year | Repayment | Year when loan is expected to be repaid |
        | Investment style | Investment style | Core, Enhanced, etc. - determines risk profile |
        | 32_PROPERTY COUNTRY | Country | Country where the underlying property is located |
        | 14 Floating Fixed rate | Fixed vs. Floating | Type of interest rate applied to the loan |
        | Exposure cum | Loan commitment EUR | Total committed amount in EUR |
        | Current loan amount cum | Loan amount drawn EUR | Amount currently drawn in EUR |
        | Open commitment cum | Loan amount undrawn/ repaid EUR | Remaining available amount in EUR |
        | Main Sector | Sector | Property sector (Office, Retail, etc.) |
        | Remaining life | Remaining loan term if fully extended | Time until maturity including extensions |
        | ARE CREL Grade | Credit Rating | Internal credit rating |
        | Interest rate / Coupon p.a. (%) | Current coupon (p.a.) | Annual interest rate |
        | LTV (%) | Latest LTV | Loan-to-Value ratio |
        | DSCR | DSCR (current) T12 capped | Debt Service Coverage Ratio |
        | ICR | ICR (current) T12 capped | Interest Coverage Ratio |
        | Net Debt Yield (%) | Debt yield (current) | Net Operating Income / Loan Amount |
        
        **PPRE US Column Mapping:**
        
        | Original Column | Standardized Column | Description |
        |----------------|---------------------|-------------|
        | Asset Manager | Asset manager | Entity managing the asset |
        | LOAN_NAME | Investment name | Primary identifier of the loan |
        | OFV | Loan commitment (loan CCY) | Total committed amount in original currency |
        | CURR_PRIN_BAL | Loan amount drawn (loan CCY) | Amount currently drawn in original currency |
        | CONSTRUCTION | Development | Indicates if the loan is for construction |
        | PROPERTY_TYPE | Sector | Property sector (Office, Retail, etc.) |
        | INTEREST_TYPE | Fixed vs. Floating | Type of interest rate applied to the loan |
        | LETTER_RATING | Credit Rating | Internal credit rating |
        | CURR_INT_RATE | Current coupon (p.a.) | Annual interest rate |
        | CURR_INT_LTV | Latest LTV | Loan-to-Value ratio |
        | NRM_DSCR | DSCR (current) T12 capped | Debt Service Coverage Ratio |
        | Watchlist | Watchlist | Risk monitoring flag |
        | MAT_DATE | Used for Remaining loan term | Maturity date for term calculation |
        """)
        
    with tabs[4]:
        st.markdown("""
        ### Detailed Data Transformation Steps
        
        **Data Cleaning Transformations:**
        
        1. **Missing Data Handling**
           - Text columns with hierarchical structure: Forward-fill to propagate parent values to children
           - Missing numeric values: Kept as NULL to avoid incorrect calculations
           - Empty strings: Converted to NULL for consistency
        
        2. **Data Type Conversion**
           - Percentage values: Converted from string format (with % sign) to decimal (divided by 100)
           - Currency values: Parsed removing thousand separators and converting to float
           - Dates: Parsed and converted to datetime objects for term calculations
        
        3. **Value Standardization**
           - Country codes: Standardized (e.g., 'GB' -> 'UK')
           - Property types: Mapped to consistent categories (e.g., 'MULTIFAM' -> 'Residential')
           - Fixed/Floating values: Standardized terminology
           - Watchlist values: Red/Yellow/Green traffic light system converted to Yes/No flags
        
        **Data Enhancement Transformations:**
        
        1. **Derived Metrics Calculation**
           - Undrawn amount: Commitment - Drawn amount
           - EUR values: Converted from local currency using exchange rate
           - Remaining terms: Calculated as (maturity_date - reporting_date) / 365 in years
        
        2. **Risk Classification**
           - LTV categorization into risk bands (<55%, 55-65%, 65-75%, >75%)
           - Development flag based on Investment style or CONSTRUCTION field
           - Credit Rating preserved from source or mapped to standard scale
        
        3. **Portfolio Analysis Metrics**
           - Weighted metrics: Key ratios weighted by loan amount for portfolio-level analysis
           - Currency normalized values: All monetary values available in both original currency and EUR
        
        **Integration Transformations:**
        
        1. **Schema Alignment**
           - Column standardization across different sources
           - Missing columns added with NULL values
           - Data types harmonized for safe concatenation
        
        2. **Deduplication**
           - Identifying key columns for duplicate detection
           - Preserving most recent/complete record when duplicates found
           - Template updates avoid duplicate entries using key-based matching
        """)
        
    with tabs[5]:
        st.markdown("""
        ### Key Calculations
        
        **LTV and Risk Classifications:**
        - **LTV (Loan-to-Value)**: Measures the loan amount relative to the property value
          * Original values from source files are preserved in 'Latest LTV' column
          * LTV values are categorized into risk bands:
            - 'LTV < 55%': Values < 0.55 (lowest risk)
            - '55% ≤ LTV < 65%': Values between 0.55 and 0.65 (moderate risk)
            - '65% ≤ LTV < 75%': Values between 0.65 and 0.75 (elevated risk)
            - '75% < LTV': Values >= 0.75 (highest risk)
        
        **Financial Performance Metrics:**
        - **DSCR (Debt Service Coverage Ratio)**: NOI / Debt Service
          * Measures ability to cover debt payments from property income
          * Higher values indicate stronger financial performance
          * T12 capped indicates trailing 12 months with cap rate application
        
        - **ICR (Interest Coverage Ratio)**: NOI / Interest Expense
          * For US data: Calculated as NRM_NOI / (Current coupon × Loan amount drawn)
          * Measures ability to cover interest payments from property income
          * Higher values indicate stronger performance
        
        - **Debt Yield**: NOI / Loan Amount
          * For US data: Calculated as NRM_NOI / Loan amount drawn
          * Direct measure of return on debt investment
          * Not affected by interest rates unlike DSCR and ICR
        
        **Weighted Portfolio Metrics:**
        - **Weighted Calculations** - Enables portfolio-level analysis:
          * 'Current coupon (p.a.) weighted' = Current coupon × Loan amount drawn EUR
          * 'Latest LTV Weighted' = Latest LTV × Loan amount drawn EUR
          * 'DSCR (current) Weighted' = DSCR × Loan amount drawn EUR
          * 'ICR (current) Weighted' = ICR × Loan commitment EUR
          * 'Debt yield (current) Weighted' = Debt yield × Loan amount drawn EUR
          * These weighted values can be summed and divided by total loan amount to get portfolio average
        
        **Currency Conversions:**
        - **USD to EUR Conversion**: 
          * Exchange rate retrieved from European Central Bank API (primary)
          * Fallback to Yahoo Finance API if ECB fails
          * Final fallback to fixed rate of 0.92 if both APIs fail
          * All monetary values maintained in both original currency and EUR
        
        **Time Calculations:**
        - **Remaining Loan Term**: 
          * Without extension: (Maturity Date - Reporting Date)/365 in years
          * With extension: Same as without extension for US, may differ for EU based on data
        """)
        
    with tabs[6]:
        st.markdown("""
        ### Technical Implementation Details
        
        **File Processing Techniques:**
        
        1. **EU Delta Master File Handling**
           - Flexible header detection algorithm to accommodate varying file structures
           - Forward-fill mechanism for hierarchical data where parent information applies to children
           - Intelligent mapping of columns based on names rather than fixed positions
        
        2. **US AZL File Handling**
           - Fixed header structure with standardized column detection
           - Specific data cleaning operations for US-formatted data
           - Exchange rate application for currency conversion
        
        3. **Watchlist Integration**
           - Sheet name determination based on year and quarter (e.g., "2023Q4")
           - Header row selection at index 6 (7th row) based on file standard
           - Left join with main data to preserve all loans even without watchlist entries
        
        **Data Concatenation Safety:**
        
        1. **Column Type Consistency**
           - Pre-processing of DataFrames to ensure consistent column types
           - Empty DataFrame with all columns created as reference
           - Each DataFrame aligned to this structure before concatenation
        
        2. **Warning Prevention**
           - Explicit handling of empty/NA entries to prevent deprecation warnings
           - Axis specification in concat operations for clarity
           - Safe handling of different data sources with varying columns
        
        **Excel Output Formatting:**
        
        1. **Beautiful Formatting Without Tables**
           - Direct cell styling instead of Excel Table objects to prevent corruption
           - Alternating row colors (white/light green) for readability
           - Header styling with blue background and white text
        
        2. **Number Formatting**
           - Currency values: '#,##0.00' format with thousand separators and 2 decimals
           - Percentages: '0.00%' format with percentage sign and 2 decimals
           - Ratios and metrics: '0.00' format with 2 decimals
        
        3. **Template Integration**
           - Safe appending to existing templates with Data_hub sheet
           - Duplicate prevention using key columns (Investment name, Asset manager, Reporting Date)
           - Preservation of all sheets and data in the template
           - Two-stage formatting: basic append followed by beautification
        """)

def main():
    st.set_page_config(page_title="Delta Master Processor", page_icon="📊", layout="wide")
    
    st.title("Delta Master Processor")
    
    tab1, tab2, tab3 = st.tabs(["Process Data", "Documentation", "Help"])
    
    with tab1:
        st.header("Process Delta Master Data")
        
        col1, col2 = st.columns(2)
        
        with col1:
            asset_manager = st.selectbox("Select Asset Manager", 
                                        options=["PPRE EU", "PPRE US", "PPRE EU + PPRE US"],
                                        help="Choose which asset manager's data to process")
        
        with col2:
            # Use a date picker instead of text input
            today = datetime.today()
            last_day_of_previous_quarter = datetime(
                today.year - (1 if today.month < 4 else 0),
                12 if today.month < 4 else (3 if today.month < 7 else 6 if today.month < 10 else 9),
                31 if today.month < 4 or 7 <= today.month < 10 else 30
            )
            
            reporting_date_obj = st.date_input(
                "Reporting Date",
                value=last_day_of_previous_quarter,
                help="Select the reporting date"
            )
            
            # Convert the date object to the required string format DD.MM.YYYY
            reporting_date = reporting_date_obj.strftime("%d.%m.%Y")
            st.write(f"Formatted date: {reporting_date}")
        
        # File upload section
        st.subheader("Upload Files")
        
        if asset_manager == "PPRE EU":
            input_files = st.file_uploader("Upload Delta Master Files", 
                                         accept_multiple_files=True,
                                         type=["xlsx", "xls"],
                                         help="Upload files containing Delta Master data. Files with 'Delta' in the name will be auto-detected.")
            
            watchlist_file = st.file_uploader("Upload Watchlist File", 
                                            type=["xlsx", "xls"],
                                            help="Upload the watchlist file with traffic light data")
        elif asset_manager == "PPRE US":
            input_files = st.file_uploader("Upload US Data Files", 
                                         accept_multiple_files=True,
                                         type=["xlsx", "xls"],
                                         help="Upload files containing US data. Files with 'AZL' in the name will be auto-detected.")
            watchlist_file = None
        else:  # Combined mode
            input_files = st.file_uploader("Upload All Data Files", 
                                         accept_multiple_files=True,
                                         type=["xlsx", "xls"],
                                         help="Upload all files for processing. Files with 'Delta' will be processed as EU files, files with 'AZL' as US files.")
            
            watchlist_file = st.file_uploader("Upload Watchlist File for EU Data", 
                                            type=["xlsx", "xls"],
                                            help="Upload the watchlist file with traffic light data for PPRE EU files")
        
        template_file = st.file_uploader("Upload Template File (Optional)", 
                                       type=["xlsx", "xls"],
                                       help="Upload a template file with 'Data_hub' sheet to append data to")
        
        # Display pandas version info
        st.info(f"Using pandas version: {pd.__version__}")
        
        # Process button
        if st.button("Process Data"):
            if not input_files:
                st.error("Please upload at least one input file.")
            else:
                try:
                    with st.spinner("Processing data..."):
                        # If PPRE EU is selected but no watchlist file is uploaded
                        if (asset_manager == "PPRE EU" or asset_manager == "PPRE EU + PPRE US") and watchlist_file is None:
                            st.warning("No watchlist file uploaded. Proceeding without watchlist integration for EU data.")
                        
                        # Process the files
                        result_df, modified_template = process_files(
                            asset_manager, 
                            reporting_date, 
                            input_files, 
                            watchlist_file,
                            template_file
                        )
                        
                        if result_df is not None:
                            # Display preview
                            st.subheader("Data Preview")
                            st.dataframe(result_df.head(10))
                            
                            # Download buttons
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                # Individual file downloads
                                st.subheader("Download Individual Files")
                                
                                # Download processed data as CSV
                                csv = result_df.to_csv(index=False)
                                b64 = base64.b64encode(csv.encode()).decode()
                                href = f'<a href="data:file/csv;base64,{b64}" download="processed_data.csv">Download Processed Data (CSV)</a>'
                                st.markdown(href, unsafe_allow_html=True)
                                
                                # Create and download beautifully formatted Excel
                                excel_data = create_excel_output(result_df)
                                if excel_data:
                                    b64_excel = base64.b64encode(excel_data.getvalue()).decode()
                                    href_excel = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_excel}" download="processed_data.xlsx">Download Processed Data (Excel)</a>'
                                    st.markdown(href_excel, unsafe_allow_html=True)
                                
                                # If template was provided and modified
                                if modified_template is not None:
                                    template_data = modified_template.getvalue()
                                    b64_template = base64.b64encode(template_data).decode()
                                    href_template = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_template}" download="updated_template.xlsx">Download Updated Template</a>'
                                    st.markdown(href_template, unsafe_allow_html=True)
                            
                            with col2:
                                # Download all as ZIP
                                st.subheader("Download All Files")
                                zip_buffer = create_download_zip(result_df, modified_template)
                                if zip_buffer:
                                    b64_zip = base64.b64encode(zip_buffer.getvalue()).decode()
                                    today_date = datetime.today().strftime("%Y-%m-%d")
                                    download_filename = f"delta_master_outputs_{today_date}.zip"
                                    href_zip = f'<a href="data:application/zip;base64,{b64_zip}" download="{download_filename}">Download All Files as ZIP</a>'
                                    st.markdown(href_zip, unsafe_allow_html=True)
                                    st.write("This ZIP file contains all output files in one download.")
                
                except Exception as e:
                    st.error(f"An error occurred during processing: {str(e)}")
                    import traceback
                    st.error(traceback.format_exc())
    
    with tab2:
        show_enhanced_documentation()
    
    with tab3:
        st.header("Help & Troubleshooting")
        
        st.markdown("""
        ### Common Issues and Solutions
        
        **File Format Issues**
        - Ensure Excel files are properly formatted and not corrupted
        - For PPRE EU files, the application will automatically detect the header row
        - For PPRE US files, ensure data is in the first sheet with headers in row 1
        
        **Missing Columns**
        - The application will warn about missing columns and create empty ones when needed
        - Check the Documentation tab for expected column mappings
        
        **Watchlist Integration**
        - For PPRE EU, watchlist file should have sheets named by year and quarter (e.g., "2023Q4")
        - Headers in watchlist file should be in row 7
        - First column should contain deal names that match "Investment name" values
        - Second column should contain traffic light values (Green, Yellow, Red)
        
        **Template File Requirements**
        - Must contain a sheet named "Data_hub"
        - Should have columns matching the standardized output structure
        - The application will only append new records based on key fields
        
        **Exchange Rate Issues**
        - The application attempts to get current exchange rates from European Central Bank
        - If connection fails, it will use Yahoo Finance or fall back to a fixed rate
        - You can see the rate being used in the process log
        
        **Combined Processing Mode**
        - When using "PPRE EU + PPRE US" mode, files are auto-categorized by name
        - Files with "Delta" in the name are processed as PPRE EU files
        - Files with "AZL" in the name are processed as PPRE US files
        - Other files are tried with both processors
        
        **Download Options**
        - Individual files can be downloaded separately (CSV, Excel, Template)
        - The "Download All Files as ZIP" option packages all outputs into a single ZIP file
        - The ZIP includes the processed data in both CSV and Excel formats, plus the updated template
        """)
        
        st.info("For further assistance, please contact the developer.")

if __name__ == "__main__":
    main()
